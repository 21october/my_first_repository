from urllib.request import urlopen
import requests
from bs4 import BeautifulSoup
import urllib.request
import openpyxl

#https://fr.sandro-paris.com/fr/femme/manteaux/manteau-en-drap-de-laine/SFPOU00489.html?dwvar_SFPOU00489_color=G023#start=1

#python3 -m streamlit run /Users/youkyung/temp/my_first_repository/my_first_repository/strlit.py
# import streamlit as st
# import pandas as pd
# form = st.form(key="my_form")
# url = form.text_input(label="Url을 입력하세요. 이미지와 상품정보가 다운로드 됩니다.")
# submit_button = form.form_submit_button(label='Submit')
# print(url)


# 엑셀파일 만들기
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["이름","가격"])

#웹사이트 정보 담기
url = input("URL 주소를 입력하세요.")
response = requests.get(url)
html = response.text
soup = BeautifulSoup(html, 'html.parser')

#이름 정보 가져와서 엑셀에 기록
name = soup.select_one("#title").text
ws.cell(row=2, column=1).value = name

#가격 정보 가져와서 엑셀에 기록
price = soup.select_one("span.price-sales").text.strip()
ws.cell(row=2, column=2).value = price

#엑셀에 기록한 내용 저장하기
wb.save(f"/Users/youkyung/miniproject/{name}.xlsx")

#이미지 정보 가져오기
images = soup.select("img",attrs={"class":"productthumbnail lazyload loaded"})

list = []
for img in images:
    if "data-hires" in img.attrs:
        imgUrl = img["data-hires"]
        list.append(imgUrl)
    else:
        pass
    

n = 1
for i in list:
    with urlopen(i) as f:
        with open("/Users/youkyung/miniproject/"+name+str(n)+'.jpg','wb') as h:
            img = f.read()
            h.write(img)
    n += 1

print("작업이 끝났습니다.")
