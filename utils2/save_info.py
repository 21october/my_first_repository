from urllib.request import urlopen
import requests
from bs4 import BeautifulSoup
import urllib.request
import openpyxl

class Sandro:
    def get_info(self, page_url):
        response = requests.get(page_url)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        name = soup.select_one("#title").text
        price = soup.select_one("span.price-sales").text.strip()
        images = soup.select("img", attrs={"class": "productthumbnail lazyload loaded"})

        return {'name': name, 'price': price, 'images': images}

    def save_info(self, info):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["이름", "가격"])
        ws.cell(row=2, column=1).value = info['name']
        ws.cell(row=2, column=2).value = info['price']

        wb.save(f"./results/{info['name']}.xlsx")

    def save_images(self, info):
        img_list =[]
        for img in info['images']:
            if "data-hires" in img.attrs:
                img_url = img["data-hires"]
                img_list.append(img_url)
            else:
                pass
        
        for index, url in enumerate(img_list):
            with urlopen(url) as f:
                with open(f"./results/{info['name']}_{str(index)}.jpg", 'wb') as h:
                    img = f.read()
                    h.write(img)
        