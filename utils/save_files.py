import requests
import openpyxl

from urllib.request import urlopen
from bs4 import BeautifulSoup


class Cup:
    def __init__(self):
        self.is_empty = True

    def drink(self):
        self.is_empty = True

    def refill(self):
        self.is_empty = False
        pass


class GoodClass:
    def get_information(self, page_url):
        response = requests.get(page_url)
        # print(response.text)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        name = soup.select_one("#title").text
        price = soup.select_one("span.price-sales").text.strip()
        images = soup.select("img", attrs={"class": "productthumbnail lazyload loaded"})

        return {'name': name, 'price': price, 'images': images}

    def save_info(self, info):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["이름", "가격"])
        ws.cell(row=2, column=1).value = info['name']
        ws.cell(row=2, column=2).value = info['price']

        wb.save(f"results/{info['name']}.xlsx")

    def save_images(self, info):
        image_array = []

        for img in info['images']:
            if "data-hires" in img.attrs:
                img_url = img["data-hires"]
                image_array.append(img_url)
            else:
                pass

        for index, url in enumerate(image_array):
            with urlopen(url) as f:
                with open(f"results/{info['name']}_{str(index)}.jpg", 'wb') as h:
                    img = f.read()
                    h.write(img)

